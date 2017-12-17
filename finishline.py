#-*- encoding:utf8 -*-
import os
import time
from selenium import webdriver
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
from openpyxl import load_workbook, Workbook

# NEW ARRIVALS 들어가는 함수
def NEW_ARRIVALS():
    subTitle = ['Men', 'Women', 'Boys', 'Girls']
    Url = ['http://www.finishline.com/store/men/_/N-1hmpde0Z1bno91t?mnid=men&Ns=sku.daysAvailable%7C0&isFilter=true&sort=sort%3Anew%20arrivals%0A%20',
           'http://www.finishline.com/store/women/_/N-1rmncopZ1bno91t?mnid=women&Ns=sku.daysAvailable%7C0&isFilter=true&sort=sort%3Anew%20arrivals%0A%20',
           'http://www.finishline.com/store/boys/_/N-1ad6v8pZ1bno91t?mnid=boys&Ns=sku.daysAvailable%7C0&isFilter=true&isFilter=true',
           'http://www.finishline.com/store/girls/_/N-goias6Z1bno91t?mnid=girls&Ns=sku.daysAvailable%7C0&isFilter=true&sort=sort%3Anew%20arrivals%0A%20']

    rowNum = 2
    imgNum = 1
    for i in range(0, 4): # 브랜드 루프
        while True:
            answer = input("NEW ARRIVALS-GENDER-" + subTitle[i] + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y": break
            elif answer == "n": break
            else: continue

        if answer == 'n' : continue

        elif answer == 'y' :
            if not os.path.isdir("01 " + Title[0]):
                os.mkdir("01 " + Title[0])

            wb = Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value="공급사(경로)")  # 공급사(경로)
            ws.cell(row=1, column=2, value="상품명")  # 상품명
            ws.cell(row=1, column=3, value="공급가")  # 공급가
            ws.cell(row=1, column=4, value="공급사 상품명")  # 공급사 상품명
            ws.cell(row=1, column=5, value="옵션입력")  # 옵션입력
            ws.cell(row=1, column=7, value="기타")  # 기타

            wb.save("01 " + Title[0] + "/#" + Title[0] + ".xlsx")

            driver.get(Url[i])

            bs4 = BeautifulSoup(driver.page_source, "html.parser")
            List1 = bs4.find_all('a', {"class": "paginationLink"})
            page = int(len(List1) / 2) - 1

            if len(List1) == 0:

                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                List = bs4.findAll('li', {'class': 'noOnModelImage'})

                for k in range(0, len(List)):  # 상품 루프
                    driver.get("http://www.finishline.com" + List[k].find('div').find('a')['href'])

                    # 공급사(경로)
                    elem1 = driver.find_element_by_id('breadcrumbs').text

                    # 상품명
                    elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                    # 공급가
                    try:
                        elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                    except:
                        elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                    # 공급사 상품명
                    elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                    # 옵션입력(색상)
                    elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                    # 옵션입력(사이즈)
                    elem6 = driver.find_element_by_id('productSizes').text
                    elem6 = elem6.replace(' ', '/')

                    # 옵션입력 합침
                    plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                    # 기타
                    elem7 = driver.find_element_by_id('productDescription').text

                    ws = wb.active

                    ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                    ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                    ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                    ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                    ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                    ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                    wb.save("01 " + Title[0] + "/#" + Title[0] + ".xlsx")

                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List4 = bs4.findAll('div', {'id': 'alt'})

                    for l in range(0, len(List4)):
                        temp = List4[l].find('div').find('img')["src"]
                        imgurl = temp.replace('Thumbnail', 'Main')
                        imgurl = imgurl.replace('\n', '')
                        urlretrieve(imgurl,
                                    '01 NEW ARRIVALS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                        l + 1).rjust(2, '0') + ".jpg")

                    rowNum = rowNum + 1
                    imgNum = imgNum + 1

                    driver.back()

            else:

                for j in range(0, page):  # 페이지 루프
                    driver.get("http://www.finishline.com" + List1[j]['href'])

                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List = bs4.findAll('li', {'class': 'noOnModelImage'})

                    for k in range(0, len(List)):  # 상품 루프
                        driver.get("http://www.finishline.com" + List[k].find('div').find('a')['href'])

                        # 공급사(경로)
                        elem1 = driver.find_element_by_id('breadcrumbs').text

                        # 상품명
                        elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                        # 공급가
                        try:
                            elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                        except:
                            elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                        # 공급사 상품명
                        elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                        # 옵션입력(색상)
                        elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                        # 옵션입력(사이즈)
                        elem6 = driver.find_element_by_id('productSizes').text
                        elem6 = elem6.replace(' ', '/')

                        # 옵션입력 합침
                        plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                        # 기타
                        elem7 = driver.find_element_by_id('productDescription').text

                        ws = wb.active

                        ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                        ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                        ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                        ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                        ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                        ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                        wb.save("01 " + Title[0] + "/#" + Title[0] + ".xlsx")

                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                        List4 = bs4.findAll('div', {'id': 'alt'})

                        for l in range(0, len(List4)):
                            temp = List4[l].find('div').find('img')["src"]
                            imgurl = temp.replace('Thumbnail', 'Main')
                            imgurl = imgurl.replace('\n', '')
                            urlretrieve(imgurl,
                                        '01 NEW ARRIVALS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                            l + 1).rjust(2, '0') + ".jpg")

                        rowNum = rowNum + 1
                        imgNum = imgNum + 1

                        driver.back()

# FAN GEAR 들어가는 함수
def FAN_GEAR():
    subTitle = ['COLLEGE', 'NFL', 'NBA', 'MLB']
    Url = []

    bs4 = BeautifulSoup(driver.page_source, "html.parser")
    List = bs4.find('ul', class_='medium-block-grid-5').findAll('li')

    for i in range(0, 4):
        Url.append("http://www.finishline.com" + List[i].find('a')['href'])

    rowNum = 2
    imgNum = 1
    for i in range(0, 4):
        while True:
            answer = input("FAN GEAR-" + subTitle[i] + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y": break
            elif answer == "n": break
            else: continue

        if answer == 'n' : continue

        elif answer == 'y' :
            driver.get(Url[i])

            for j in range(1, 7):
                if j != 1:
                    driver.find_element_by_xpath('//*[@id="auxNav"]/div/div[2]/ul/li[' + str(j) + ']/a/span[2]').click()

                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                List = bs4.find('ul', class_ = 'leaugeContent active').findAll('li')

                for k in range(0, len(List)):
                    data = List[k].find('a')

                    while True:
                        answer = input("FAN GEAR-" + subTitle[i] + "-" + data.text + "의 데이터를 추출하시겠습니까? y/n : ")
                        if answer == "y": break
                        elif answer == "n": break
                        else: continue

                    if answer == 'n': continue

                    elif answer == 'y':
                        try: driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
                        except: driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

                        if not os.path.isdir("02 " + Title[1]):
                            os.mkdir("02 " + Title[1])

                        wb = Workbook()
                        ws = wb.active

                        ws.cell(row=1, column=1, value="공급사(경로)")  # 공급사(경로)
                        ws.cell(row=1, column=2, value="상품명")  # 상품명
                        ws.cell(row=1, column=3, value="상품명(관리용)")  # 상품명(관리용)
                        ws.cell(row=1, column=4, value="공급사 상품명")  # 공급사 상품명
                        ws.cell(row=1, column=5, value="공급가")  # 공급가
                        ws.cell(row=1, column=6, value="변동가")  # 변동가
                        ws.cell(row=1, column=7, value="공급가변환($->\)")  # 공급가변환($->\)
                        ws.cell(row=1, column=8, value="옵션입력")  # 옵션입력
                        ws.cell(row=1, column=9, value="등록일")  # 등록일
                        ws.cell(row=1, column=10, value="품절")  # 품절
                        ws.cell(row=1, column=11, value="기타")  # 기타

                        wb.save("02 " + Title[1] + "/#" + Title[1] + ".xlsx")
                        driver.get("http://www.finishline.com" + data['href'])

                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                        List1 = bs4.find_all('a', {"class": "paginationLink"})
                        page = int(len(List1) / 2) - 1

                        if len(List1) == 0:
                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                            List = bs4.findAll('li', {'class': 'noOnModelImage'})

                            for p in range(0, len(List)):  # 상품 루프
                                driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                # 공급사(경로)
                                elem1 = driver.find_element_by_id('breadcrumbs').text

                                # 상품명
                                elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                # 공급가
                                try:
                                    elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                except:
                                    elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                # 공급사 상품명
                                elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                # 옵션입력(색상)
                                elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                # 옵션입력(사이즈)
                                elem6 = driver.find_element_by_id('productSizes').text
                                elem6 = elem6.replace(' ', '/')

                                # 옵션입력 합침
                                plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                # 기타
                                elem7 = driver.find_element_by_id('productDescription').text

                                ws = wb.active

                                ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                wb.save("02 " + Title[1] + "/#" + Title[1] + ".xlsx")

                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                List4 = bs4.findAll('div', {'id': 'alt'})

                                for l in range(0, len(List4)):
                                    temp = List4[l].find('div').find('img')["src"]
                                    imgurl = temp.replace('Thumbnail', 'Main')
                                    imgurl = imgurl.replace('\n', '')
                                    urlretrieve(imgurl,
                                                '02 FAN GEAR' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                    l + 1).rjust(2, '0') + ".jpg")

                                rowNum = rowNum + 1
                                imgNum = imgNum + 1

                                driver.back()

                        else :
                            for o in range(0, page):  # 페이지 루프
                                driver.get("http://www.finishline.com" + List1[o]['href'])

                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                for p in range(0, len(List)):  # 상품 루프
                                    driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                    # 공급사(경로)
                                    elem1 = driver.find_element_by_id('breadcrumbs').text

                                    # 상품명
                                    elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                    # 공급가
                                    try:
                                        elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                    except:
                                        elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                    # 공급사 상품명
                                    elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                    # 옵션입력(색상)
                                    elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                    # 옵션입력(사이즈)
                                    elem6 = driver.find_element_by_id('productSizes').text
                                    elem6 = elem6.replace(' ', '/')

                                    # 옵션입력 합침
                                    plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                    # 기타
                                    elem7 = driver.find_element_by_id('productDescription').text

                                    ws = wb.active

                                    ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                    ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                    ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                    ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                    ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                    ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                    wb.save("02 " + Title[1] + "/#" + Title[1] + ".xlsx")

                                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                    List4 = bs4.findAll('div', {'id': 'alt'})

                                    for l in range(0, len(List4)):
                                        temp = List4[l].find('div').find('img')["src"]
                                        imgurl = temp.replace('Thumbnail', 'Main')
                                        imgurl = imgurl.replace('\n', '')
                                        urlretrieve(imgurl,
                                                    '02 FAN GEAR' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                        l + 1).rjust(2, '0') + ".jpg")

                                    rowNum = rowNum + 1
                                    imgNum = imgNum + 1

                                    driver.back()

                        driver.get(Url[i])

# BRANDS 들어가는 함수
def BRANDS():
    Url = "http://www.finishline.com"

    bs4 = BeautifulSoup(driver.page_source, "html.parser")
    List = bs4.find('ul', class_='subCategory').findAll('li')

    rowNum = 2
    imgNum = 1
    for i in range(0, len(List) + 2):
        if i < 2: data = driver.find_element_by_xpath('//*[@id="grid"]/div[1]/div/section/div[6]/div[1]/li[' + str(i + 1) + ']/a')
        else: data = driver.find_element_by_xpath('//*[@id="grid"]/div[1]/div/section/div[6]/div[1]/ul/li[' + str(i - 1) + ']/a')
        MWK = data.text

        while True:
            answer = input("BRANDS-" + MWK + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y": break
            elif answer == "n": break
            else: continue

        if answer == 'n': continue

        elif answer == 'y':
            data.click()

            if not os.path.isdir("03 " + Title[2]):
                os.mkdir("03 " + Title[2])

            wb = Workbook()
            ws = wb.active

            ws.cell(row=1, column=1, value="공급사(경로)")  # 공급사(경로)
            ws.cell(row=1, column=2, value="상품명")  # 상품명
            ws.cell(row=1, column=3, value="상품명(관리용)")  # 상품명(관리용)
            ws.cell(row=1, column=4, value="공급사 상품명")  # 공급사 상품명
            ws.cell(row=1, column=5, value="공급가")  # 공급가
            ws.cell(row=1, column=6, value="변동가")  # 변동가
            ws.cell(row=1, column=7, value="공급가변환($->\)")  # 공급가변환($->\)
            ws.cell(row=1, column=8, value="옵션입력")  # 옵션입력
            ws.cell(row=1, column=9, value="등록일")  # 등록일
            ws.cell(row=1, column=10, value="품절")  # 품절
            ws.cell(row=1, column=11, value="기타")  # 기타

            wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

            ch = False
            try: # 1
                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                L01 = bs4.findAll('span', class_='linkCombination')

                for j in range(0, len(L01)): # men, women, kid
                    title = L01[j].find('span', class_='firstHead').text.strip()
                    # print(title)
                    if (title == 'Mens') | (title == 'Women') | (title == 'Kids'):
                        ch = True
                        while True:
                            answer = input("BRANDS-" + MWK + '-' + title + "의 데이터를 추출하시겠습니까? y/n : ")
                            if answer == "y": break
                            elif answer == "n": break
                            else: continue

                        if answer == 'n': continue

                        elif answer == 'y':
                            L02 = L01[j].findAll('a')
                            for k in range(1, len(L02)):
                                T02 = L02[k].text
                                while True:
                                    answer = input("BRANDS-" + MWK + '-' + title + '-' + T02 + "의 데이터를 추출하시겠습니까? y/n : ")
                                    if answer == "y": break
                                    elif answer == "n": break
                                    else: continue

                                if answer == 'n': continue

                                elif answer == 'y':
                                    driver.get(Url + L02[k]['href'])

                                    if (T02 == "Boys' Shoes") | (T02 == "Girls' Shoes"): # kid 예외사항
                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                        L03 = bs4.find('div', class_='show-for-medium-up medium-3 large-3 xlarge-2 columns products-sidebar filters').findAll('a', class_='sub featured first-child')

                                        for l in range(0, len(L03)):
                                            while True:
                                                answer = input("BRANDS-" + MWK + '-' + title + '-' + T02 + '-'+ L03[l].text + "의 데이터를 추출하시겠습니까? y/n : ")
                                                if answer == "y": break
                                                elif answer == "n": break
                                                else: continue

                                            if answer == 'n': continue

                                            elif answer == 'y':
                                                driver.get(Url + L03[l]['href'])
                                                try: driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
                                                except: driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

                                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                List1 = bs4.find_all('a', {"class": "paginationLink"})
                                                page = int(len(List1) / 2) - 1

                                                if len(List1) == 0:
                                                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                    List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                                    for p in range(0, len(List)):  # 상품 루프
                                                        driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                                        # 공급사(경로)
                                                        elem1 = driver.find_element_by_id('breadcrumbs').text

                                                        # 상품명
                                                        elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                                        # 공급가
                                                        try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                                        except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                                        # 공급사 상품명
                                                        elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                                        # 옵션입력(색상)
                                                        elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                                        # 옵션입력(사이즈)
                                                        elem6 = driver.find_element_by_id('productSizes').text
                                                        elem6 = elem6.replace(' ', '/')

                                                        # 옵션입력 합침
                                                        plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                                        # 기타
                                                        elem7 = driver.find_element_by_id('productDescription').text

                                                        ws = wb.active

                                                        ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                                        ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                                        ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                                        ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                                        ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                                        ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                                        wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                        List4 = bs4.findAll('div', {'id': 'alt'})

                                                        for l in range(0, len(List4)):
                                                            temp = List4[l].find('div').find('img')["src"]
                                                            imgurl = temp.replace('Thumbnail', 'Main')
                                                            imgurl = imgurl.replace('\n', '')
                                                            urlretrieve(imgurl,
                                                                        '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                                            l + 1).rjust(2, '0') + ".jpg")

                                                        rowNum = rowNum + 1
                                                        imgNum = imgNum + 1

                                                        driver.back()

                                                else :
                                                    for o in range(0, page):  # 페이지 루프
                                                        driver.get("http://www.finishline.com" + List1[o]['href'])

                                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                        List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                                        for p in range(0, len(List)):  # 상품 루프
                                                            driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                                            # 공급사(경로)
                                                            elem1 = driver.find_element_by_id('breadcrumbs').text

                                                            # 상품명
                                                            elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                                            # 공급가
                                                            try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                                            except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                                            # 공급사 상품명
                                                            elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                                            # 옵션입력(색상)
                                                            elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                                            # 옵션입력(사이즈)
                                                            elem6 = driver.find_element_by_id('productSizes').text
                                                            elem6 = elem6.replace(' ', '/')

                                                            # 옵션입력 합침
                                                            plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                                            # 기타
                                                            elem7 = driver.find_element_by_id('productDescription').text

                                                            ws = wb.active

                                                            ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                                            ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                                            ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                                            ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                                            ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                                            ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                                            wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                            List4 = bs4.findAll('div', {'id': 'alt'})

                                                            for l in range(0, len(List4)):
                                                                temp = List4[l].find('div').find('img')["src"]
                                                                imgurl = temp.replace('Thumbnail', 'Main')
                                                                imgurl = imgurl.replace('\n', '')
                                                                urlretrieve(imgurl,
                                                                            '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                                                l + 1).rjust(2, '0') + ".jpg")

                                                            rowNum = rowNum + 1
                                                            imgNum = imgNum + 1

                                                            driver.back()

                                    else: # 그 외
                                        try: driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
                                        except: driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                        List1 = bs4.find_all('a', {"class": "paginationLink"})
                                        page = int(len(List1) / 2) - 1

                                        if len(List1) == 0:
                                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                            List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                            for p in range(0, len(List)):  # 상품 루프
                                                driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                                # 공급사(경로)
                                                elem1 = driver.find_element_by_id('breadcrumbs').text

                                                # 상품명
                                                elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                                # 공급가
                                                try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                                except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                                # 공급사 상품명
                                                elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                                # 옵션입력(색상)
                                                elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                                # 옵션입력(사이즈)
                                                elem6 = driver.find_element_by_id('productSizes').text
                                                elem6 = elem6.replace(' ', '/')

                                                # 옵션입력 합침
                                                plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                                # 기타
                                                elem7 = driver.find_element_by_id('productDescription').text

                                                ws = wb.active

                                                ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                                ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                                ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                                ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                                ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                                ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                                wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                List4 = bs4.findAll('div', {'id': 'alt'})

                                                for l in range(0, len(List4)):
                                                    temp = List4[l].find('div').find('img')["src"]
                                                    imgurl = temp.replace('Thumbnail', 'Main')
                                                    imgurl = imgurl.replace('\n', '')
                                                    urlretrieve(imgurl,
                                                                '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                                    l + 1).rjust(2, '0') + ".jpg")

                                                rowNum = rowNum + 1
                                                imgNum = imgNum + 1

                                                driver.back()

                                        else :
                                            for o in range(0, page):  # 페이지 루프
                                                driver.get("http://www.finishline.com" + List1[o]['href'])

                                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                                for p in range(0, len(List)):  # 상품 루프
                                                    driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                                    # 공급사(경로)
                                                    elem1 = driver.find_element_by_id('breadcrumbs').text

                                                    # 상품명
                                                    elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                                    # 공급가
                                                    try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                                    except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                                    # 공급사 상품명
                                                    elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                                    # 옵션입력(색상)
                                                    elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                                    # 옵션입력(사이즈)
                                                    elem6 = driver.find_element_by_id('productSizes').text
                                                    elem6 = elem6.replace(' ', '/')

                                                    # 옵션입력 합침
                                                    plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                                    # 기타
                                                    elem7 = driver.find_element_by_id('productDescription').text

                                                    ws = wb.active

                                                    ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                                    ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                                    ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                                    ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                                    ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                                    ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                                    wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                                    List4 = bs4.findAll('div', {'id': 'alt'})

                                                    for l in range(0, len(List4)):
                                                        temp = List4[l].find('div').find('img')["src"]
                                                        imgurl = temp.replace('Thumbnail', 'Main')
                                                        imgurl = imgurl.replace('\n', '')
                                                        urlretrieve(imgurl,
                                                                    '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                                        l + 1).rjust(2, '0') + ".jpg")

                                                    rowNum = rowNum + 1
                                                    imgNum = imgNum + 1

                                                    driver.back()

                                    driver.find_element_by_xpath('//*[@id="grid"]/div[1]/div/section/div[3]/div/ul/li[3]/a').click()

                if ch == False: data.click()

            except:
                ch2 = False
                try: # 2
                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List = bs4.find('ul', class_='gender').findAll('a')
    
                    for j in range(0, len(List)):
                        Link = Url + List[j]['href'].replace('isFilter=true', 'Ns=sku.daysAvailable%7C0&isFilter=true')
                        name = List[j].findAll('span')
                        name = name[1].text
    
                        if (name == 'Men') | (name == 'Women'):
                            ch2 = True
                            while True:
                                answer = input("BRANDS-" + MWK + '-' + name + "의 데이터를 추출하시겠습니까? y/n : ")
                                if answer == "y": break
                                elif answer == "n": break
                                else: continue

                            if answer == 'n': continue

                            elif answer == 'y':
                                driver.get(Link)

                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                List1 = bs4.find_all('a', {"class": "paginationLink"})
                                page = int(len(List1) / 2) - 1

                                if len(List1) == 0:
                                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                    List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                    for p in range(0, len(List)):  # 상품 루프
                                        driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                        # 공급사(경로)
                                        elem1 = driver.find_element_by_id('breadcrumbs').text

                                        # 상품명
                                        elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                        # 공급가
                                        try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                        except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                        # 공급사 상품명
                                        elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                        # 옵션입력(색상)
                                        elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                        # 옵션입력(사이즈)
                                        elem6 = driver.find_element_by_id('productSizes').text
                                        elem6 = elem6.replace(' ', '/')

                                        # 옵션입력 합침
                                        plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                        # 기타
                                        elem7 = driver.find_element_by_id('productDescription').text

                                        ws = wb.active

                                        ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                        ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                        ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                        ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                        ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                        ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                        wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                        List4 = bs4.findAll('div', {'id': 'alt'})

                                        for l in range(0, len(List4)):
                                            temp = List4[l].find('div').find('img')["src"]
                                            imgurl = temp.replace('Thumbnail', 'Main')
                                            imgurl = imgurl.replace('\n', '')
                                            urlretrieve(imgurl,
                                                        '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                            l + 1).rjust(2, '0') + ".jpg")

                                        rowNum = rowNum + 1
                                        imgNum = imgNum + 1

                                        driver.back()

                                else :
                                    for o in range(0, page):  # 페이지 루프
                                        driver.get("http://www.finishline.com" + List1[o]['href'])

                                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                        List = bs4.findAll('li', {'class': 'noOnModelImage'})

                                        for p in range(0, len(List)):  # 상품 루프
                                            driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                            # 공급사(경로)
                                            elem1 = driver.find_element_by_id('breadcrumbs').text

                                            # 상품명
                                            elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                            # 공급가
                                            try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                            except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                            # 공급사 상품명
                                            elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                            # 옵션입력(색상)
                                            elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                            # 옵션입력(사이즈)
                                            elem6 = driver.find_element_by_id('productSizes').text
                                            elem6 = elem6.replace(' ', '/')

                                            # 옵션입력 합침
                                            plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                            # 기타
                                            elem7 = driver.find_element_by_id('productDescription').text

                                            ws = wb.active

                                            ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                            ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                            ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                            ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                            ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                            ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                            wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                            List4 = bs4.findAll('div', {'id': 'alt'})

                                            for l in range(0, len(List4)):
                                                temp = List4[l].find('div').find('img')["src"]
                                                imgurl = temp.replace('Thumbnail', 'Main')
                                                imgurl = imgurl.replace('\n', '')
                                                urlretrieve(imgurl,
                                                            '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                                l + 1).rjust(2, '0') + ".jpg")

                                            rowNum = rowNum + 1
                                            imgNum = imgNum + 1

                                            driver.back()

                    if ch2 == False: data.click()

                except: # 3
                    try: driver.find_element_by_xpath('//*[@id="sortby"]/option[2]').click()
                    except: driver.find_element_by_xpath('//*[@id="fysSort"]/option[3]').click()

                    bs4 = BeautifulSoup(driver.page_source, "html.parser")
                    List1 = bs4.find_all('a', {"class": "paginationLink"})
                    page = int(len(List1) / 2) - 1

                    if len(List1) == 0:
                        bs4 = BeautifulSoup(driver.page_source, "html.parser")
                        List = bs4.findAll('li', {'class': 'noOnModelImage'})

                        for p in range(0, len(List)):  # 상품 루프
                            driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                            # 공급사(경로)
                            elem1 = driver.find_element_by_id('breadcrumbs').text

                            # 상품명
                            elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                            # 공급가
                            try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                            except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                            # 공급사 상품명
                            elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                            # 옵션입력(색상)
                            elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                            # 옵션입력(사이즈)
                            elem6 = driver.find_element_by_id('productSizes').text
                            elem6 = elem6.replace(' ', '/')

                            # 옵션입력 합침
                            plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                            # 기타
                            elem7 = driver.find_element_by_id('productDescription').text

                            ws = wb.active

                            ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                            ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                            ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                            ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                            ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                            ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                            wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                            List4 = bs4.findAll('div', {'id': 'alt'})

                            for l in range(0, len(List4)):
                                temp = List4[l].find('div').find('img')["src"]
                                imgurl = temp.replace('Thumbnail', 'Main')
                                imgurl = imgurl.replace('\n', '')
                                urlretrieve(imgurl,
                                            '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                l + 1).rjust(2, '0') + ".jpg")

                            rowNum = rowNum + 1
                            imgNum = imgNum + 1

                            driver.back()

                    else :
                        for o in range(0, page):  # 페이지 루프
                            driver.get("http://www.finishline.com" + List1[o]['href'])

                            bs4 = BeautifulSoup(driver.page_source, "html.parser")
                            List = bs4.findAll('li', {'class': 'noOnModelImage'})

                            for p in range(0, len(List)):  # 상품 루프
                                driver.get("http://www.finishline.com" + List[p].find('div').find('a')['href'])

                                # 공급사(경로)
                                elem1 = driver.find_element_by_id('breadcrumbs').text

                                # 상품명
                                elem2 = driver.find_element_by_xpath('//*[@id="title"]').text

                                # 공급가
                                try: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span[1]').text
                                except: elem3 = driver.find_element_by_xpath('//*[@id="productPrice"]/div/span')

                                # 공급사 상품명
                                elem4 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[1]').text

                                # 옵션입력(색상)
                                elem5 = driver.find_element_by_xpath('//*[@id="styleColors"]/span[2]').text

                                # 옵션입력(사이즈)
                                elem6 = driver.find_element_by_id('productSizes').text
                                elem6 = elem6.replace(' ', '/')

                                # 옵션입력 합침
                                plus = "색상 {" + elem5 + "} // 사이즈 {" + elem6 + "}"

                                # 기타
                                elem7 = driver.find_element_by_id('productDescription').text

                                ws = wb.active

                                ws.cell(row=rowNum, column=1, value=elem1)  # 공급사(경로)
                                ws.cell(row=rowNum, column=2, value=elem2)  # 상품명
                                ws.cell(row=rowNum, column=3, value=elem3)  # 공급가
                                ws.cell(row=rowNum, column=4, value=elem4)  # 공급사 상품명
                                ws.cell(row=rowNum, column=5, value=plus)  # 옵션입력
                                ws.cell(row=rowNum, column=7, value=elem7)  # 기타

                                wb.save("03 " + Title[2] + "/#" + Title[2] + ".xlsx")

                                bs4 = BeautifulSoup(driver.page_source, "html.parser")
                                List4 = bs4.findAll('div', {'id': 'alt'})

                                for l in range(0, len(List4)):
                                    temp = List4[l].find('div').find('img')["src"]
                                    imgurl = temp.replace('Thumbnail', 'Main')
                                    imgurl = imgurl.replace('\n', '')
                                    urlretrieve(imgurl,
                                                '03 BRANDS' + "/" + str(imgNum).rjust(5, '0') + "_img" + str(
                                                    l + 1).rjust(2, '0') + ".jpg")

                                rowNum = rowNum + 1
                                imgNum = imgNum + 1

                                driver.back()

            # 뒤로가기
            driver.find_element_by_xpath('//ul[@class="drp-nav-list"]/li[7]/a/span').click()

if __name__=="__main__":
    Title = ['NEW ARRIVALS', "FAN GEAR", "BRANDS"]

    url = './chromedriver'  # 드라이브가 있는 경로
    driver = webdriver.Chrome(url)
    driver.get("http://www.finishline.com")

    for i in range(0, 3):

        while True:
            answer = input(Title[i] + "의 데이터를 추출하시겠습니까? y/n : ")
            if answer == "y": break
            elif answer == "n": break
            else: continue

        if answer == 'n' : continue

        elif answer == 'y' :
            if i == 0: # NEW ARRIVALS
                driver.find_element_by_xpath('//ul[@class="drp-nav-list"]/li[1]/a/span').click()
                NEW_ARRIVALS()
            elif i == 1: # FAN GEAR
                FAN_GEAR()
            elif i == 2: # BRANDS
                driver.find_element_by_xpath('//ul[@class="drp-nav-list"]/li[7]/a/span').click()
                BRANDS()